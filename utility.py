import matplotlib
# matplotlib.use('TkAgg')

import os
import zipfile as zp
import io
import requests
import unicodedata
from openpyxl import load_workbook
# from tqdm import tqdm_notebook as tqdm
import itertools
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import re
from string import punctuation as punct
from bs4 import BeautifulSoup
from nltk.corpus import stopwords
from nltk import word_tokenize
from nltk.stem import SnowballStemmer, RSLPStemmer
from sklearn.metrics import recall_score
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.base import BaseEstimator, TransformerMixin
from joblib import dump
import plotly.graph_objects as go
import plotly.offline as pyo
from collections import Counter, defaultdict

from sklearn.model_selection._split import _BaseKFold, _RepeatedSplits
from sklearn.utils import check_random_state


def get_files_zip():
    """
    Function that lists files inside the nlp_data.zip folder.

    Returns:
        - list of paths to files.
    """
    with zp.ZipFile("./nlp_data.zip") as myzip:
        list = myzip.namelist()
    return list


def read_txt_zip(files):
    """
    Function that reads files inside the nlp_data.zip folder.

    :param files: list of strings with paths to files.

    Returns:
        - list of strings with content of corresponding files.
    """
    df_list = []
    for file in files:
        with zp.ZipFile("./nlp_data.zip") as myzip:
            with myzip.open(file) as myfile:
                df_list.append(myfile.read().decode("utf-8"))
    return df_list


# Custom Transformer that extracts columns passed as argument to its constructor
class TextCleaner(BaseEstimator, TransformerMixin):

    # Class Constructor
    def __init__(self, punctuation, stoppers, stopwords, accentuation=True, stemmer="Snowball"):
        """
            :param punctuation: list of punctuation to exclude in the text.
            :param stoppers: list of punctuation that defines the end of an excerpt.
            :param stopwords: list of stopwords to exclude.
            :accentuation: whether or not to exclude word accentuation.
            :param stemmer: "RSLP" applies RSLPStemmer('portuguese'), whereas "Snowball" applies
            SnowballStemmer('portuguese') (default). None value doesn't apply any Stemmer.
        """
        self._punctuation = punctuation
        self._stoppers = stoppers
        self._stopwords = stopwords
        if stemmer:
            if stemmer == "Snowball":
                self._stemmer = SnowballStemmer('portuguese')
            elif stemmer == "RSLP":
                self._stemmer = RSLPStemmer('portuguese')
            else:
                print("Invalid value for stemmer parameter. Default value will be used.")
                self._stemmer = SnowballStemmer('portuguese')
        else:
            self._stemmer = stemmer
        self._accentuation = accentuation


    # Return self nothing else to do here
    def fit(self, X, y=None):
        return self

    # Method that describes what we need this transformer to do
    def transform(self, X, y=None):
        """
            :param X: numpy array with texts for cleaning
        """
        url = "https://simplemaps.com/static/data/country-cities/pt/pt.csv"
        cities = pd.read_csv(io.StringIO(requests.get(url).content.decode('utf-8')))["city"].to_list()
        cities.append("lisboa")
        cities = list(map(lambda x: x.lower(), cities))
        week_days = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado',
                     'domingo']
        updates = []
        for j in X:
            # LOWERCASING
            text = j.lower()

            # REMOVING # SIGN (don't remove with punctuation as it will destroy tokens)
            text = text.replace("#", " ")

            # TOKENIZATION
            text = re.sub('\S+\$\d+( réis)*', ' #PRICE ', text)  # price token
            text = re.sub('(\d{1,2}\D\d{1,2}\D\d{2,4})|(\d{4}\D\d{1,2}\D\d{1,2})|(\d{1,2} de [a-zA-Z]+ de \d{2,4})',
                          " #DATE ", text)  # date token
            text = re.sub('[0-9]+', ' #NUMBER ', text)  # number token
            # cities token
            for i in cities:
                if i in text:
                    text = text.replace(i, ' #CITY ')
            # week days token
            for i in week_days:
                if i in text:
                    text = text.replace(i, ' #WDAY ')

            # REMOVE PUNCTUATION, SEPARATE KEEP PUNCTUATION WITH SPACES AND KEEP TOKENS #
            def repl(matchobj):
                keep_punct = list(set(punct) - (set(self._punctuation) - set(self._stoppers)) - set("#"))
                if matchobj.group(0) in keep_punct:
                    return " {} ".format(matchobj.group(0))
                elif matchobj.group(0) == "#":
                    return '#'
                else:
                    return ' '

            text = re.sub("([^a-zA-Z0-9])", repl, text)

            # REMOVE STOPWORDS
            text = " ".join(" " if word in self._stopwords else word for word in text.split())

            # REMOVE HTML TAGS
            text = BeautifulSoup(text, features="lxml").get_text()

            # APPLIES STEMMING
            if self._stemmer:
                text = " ".join(self._stemmer.stem(word) for word in text.split())

            # REMOVING ACCENTUATION
            if self._accentuation:
                text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

            updates.append(text)

        return np.array(updates)


def update_df(dataframe, list_updated):
    """
    Function that updates a DataFrame's "text" column inplace.

    :param dataframe: DataFrame to update.
    :param list_updated: list to replace the "text" column.
    """
    dataframe.update(pd.Series(list_updated, name="text"))  # Modify in place using non-NA values from another DataFrame


# def clean(text_list, punctuation, stoppers, lemmatize=None, stemmer=None):
#     """
#     Function that a receives a list of strings and preprocesses it.
#
#     :param text_list: list of strings.
#     :param punctuation: list of punctuation to exclude.
#     :param stoppers: list of punctuation that defines the end of an excerpt.
#     :param lemmatize: lemmatizer to apply if provided.
#     :param stemmer: stemmer to apply if provided.
#
#     Returns:
#         - list of cleaned strings.
#     """
#     updates = []
#     for j in range(len(text_list)):
#
#         text = text_list[j]
#
#         # LOWERCASING
#         text = text.lower()
#
#         # REMOVING # SIGN (don't remove with punctuation as it will destroy tokens)
#         text = text.replace("#", " ")
#
#         # TOKENIZATION
#         text = re.sub('\S+\$\d+( réis)*', ' #PRICE ', text)  # price token
#         text = re.sub('(\d{1,2}\D\d{1,2}\D\d{2,4})|(\d{4}\D\d{1,2}\D\d{1,2})|(\d{1,2} de [a-zA-Z]+ de \d{2,4})',
#                       " #DATE ", text)  # date token
#         text = re.sub('[0-9]+', ' #NUMBER ', text)  # number token
#         # cities token
#         url = "https://simplemaps.com/static/data/country-cities/pt/pt.csv"
#         cities = pd.read_csv(io.StringIO(requests.get(url).content.decode('utf-8')))["city"].to_list()
#         cities.append("lisboa")
#         cities = list(map(lambda x: x.lower(), cities))
#         for i in cities:
#             if i in text:
#                 text = text.replace(i, ' #CITY ')
#         # week days token
#         week_days = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado',
#                      'domingo']
#         for i in week_days:
#             if i in text:
#                 text = text.replace(i, ' #WDAY ')
#
#         # REMOVE PUNCTUATION, SEPARATE KEEP PUNCTUATION WITH SPACES AND KEEP TOKENS #
#         def repl(matchobj):
#             keep_punct = list(set(punct) - (set(punctuation) - set(stoppers)) - set("#"))
#             if matchobj.group(0) in keep_punct:
#                 return " {} ".format(matchobj.group(0))
#             elif matchobj.group(0) == "#":
#                 return '#'
#             else:
#                 return ' '
#         text = re.sub("([^a-zA-Z0-9])", repl, text)
#
#         # REMOVE STOPWORDS
#         stop = set(stopwords.words('portuguese'))
#         text = " ".join(" " if word in stop else word for word in text.split())
#
#         # REMOVE HTML TAGS
#         text = BeautifulSoup(text, features="lxml").get_text()
#
#         if lemmatize:
#             lemma = None  # Not defined
#             text = " ".join(lemma.lemmatize(word) for word in text.split())
#
#         if stemmer:
#             # stemmer = nltk.stem.RSLPStemmer()
#             # stemmer = nltk.stem.SnowballStemmer('portuguese')
#             text = " ".join(stemmer.stem(word) for word in text.split())
#
#         # REMOVING ACCENTUATION
#         text = ''.join(c for c in unicodedata.normalize('NFD', text)
#                        if unicodedata.category(c) != 'Mn')
#
#         updates.append(text)
#
#     return updates

# class ExcerptCreator(BaseEstimator, TransformerMixin):
#
#     # Class Constructor
#     def __init__(self, stoppers, size):
#         """
#             :param punctuation: list of punctuation to exclude in the text.
#             :param stoppers: list of punctuation that defines the end of an excerpt.
#             :param stopwords: list of stopwords to exclude.
#             :accentuation: whether or not to exclude word accentuation.
#             :param stemmer: "RSLP" applies RSLPStemmer('portuguese'), whereas "Snowball" applies
#             SnowballStemmer('portuguese') (default). None value doesn't apply any Stemmer.
#         """
#         self._size = size
#         self._stoppers = stoppers
#
#     # Return self nothing else to do here
#     def fit(self, X, y=None):
#         return self
#
#     # Method that describes what we need this transformer to do
#     def transform(self, X, y=None):
#         """
#             :param X: numpy array with texts for cleaning
#         """
#
#         data = []
#         for _, row in dataframe.iterrows():
#             words = word_tokenize(row["text"])
#             groups = []
#             group = []
#             for word in words:
#                 group.append(word)
#                 if (len(group) >= 485) & (group[-1] in stoppers):
#                     groups.append((row["book_id"], row["author"], " ".join(group)))
#                     group = []
#                 elif len(group) >= 600:  # Dealing with lack of punctuation
#                     groups.append((row["book_id"], row["author"], " ".join(group)))
#                     group = []
#             data += groups
#         return pd.DataFrame(data, columns=["book_id", "author", "text"])
#
#         url = "https://simplemaps.com/static/data/country-cities/pt/pt.csv"
#         cities = pd.read_csv(io.StringIO(requests.get(url).content.decode('utf-8')))["city"].to_list()
#         cities.append("lisboa")
#         cities = list(map(lambda x: x.lower(), cities))
#         week_days = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado',
#                      'domingo']
#         updates = []
#         for j in X:
#             # LOWERCASING
#             text = j.lower()
#
#             # REMOVING # SIGN (don't remove with punctuation as it will destroy tokens)
#             text = text.replace("#", " ")
#
#             # TOKENIZATION
#             text = re.sub('\S+\$\d+( réis)*', ' #PRICE ', text)  # price token
#             text = re.sub('(\d{1,2}\D\d{1,2}\D\d{2,4})|(\d{4}\D\d{1,2}\D\d{1,2})|(\d{1,2} de [a-zA-Z]+ de \d{2,4})',
#                           " #DATE ", text)  # date token
#             text = re.sub('[0-9]+', ' #NUMBER ', text)  # number token
#             # cities token
#             for i in cities:
#                 if i in text:
#                     text = text.replace(i, ' #CITY ')
#             # week days token
#             for i in week_days:
#                 if i in text:
#                     text = text.replace(i, ' #WDAY ')
#
#             # REMOVE PUNCTUATION, SEPARATE KEEP PUNCTUATION WITH SPACES AND KEEP TOKENS #
#             def repl(matchobj):
#                 keep_punct = list(set(punct) - (set(self._punctuation) - set(self._stoppers)) - set("#"))
#                 if matchobj.group(0) in keep_punct:
#                     return " {} ".format(matchobj.group(0))
#                 elif matchobj.group(0) == "#":
#                     return '#'
#                 else:
#                     return ' '
#
#             text = re.sub("([^a-zA-Z0-9])", repl, text)
#
#             # REMOVE STOPWORDS
#             text = " ".join(" " if word in self._stopwords else word for word in text.split())
#
#             # REMOVE HTML TAGS
#             text = BeautifulSoup(text, features="lxml").get_text()
#
#             # APPLIES STEMMING
#             if self._stemmer:
#                 text = " ".join(self._stemmer.stem(word) for word in text.split())
#
#             # REMOVING ACCENTUATION
#             if self._accentuation:
#                 text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
#
#             updates.append(text)
#
#         return np.array(updates)


def sample_excerpts(dataframe, stoppers, size):
    """
    Function that receives a DataFrame and creates subsets of ~500 words of each
    string in its "text" column.

    :param dataframe: DataFrame to create subsets.
    :param stoppers: list of punctuation that defines the end of an excerpt.

    Returns:
        - DataFrame object with a new text column that contains the subsets.
    """
    data = []
    for _, row in dataframe.iterrows():
        words = word_tokenize(row["text"])
        groups = []
        group = []
        for word in words:
            group.append(word)
            if size == 'short':
                if (len(group) >= 485) & (group[-1] in stoppers):
                    groups.append((row["book_id"], row["author"], " ".join(group)))
                    group = []
                elif len(group) >= 600:  # Dealing with lack of punctuation
                    groups.append((row["book_id"], row["author"], " ".join(group)))
                    group = []
            elif size == 'large':
                if (len(group) >= 980) & (group[-1] in stoppers):
                    groups.append((row["book_id"], row["author"], " ".join(group)))
                    group = []
                elif len(group) >= 1100:  # Dealing with lack of punctuation
                    groups.append((row["book_id"], row["author"], " ".join(group)))
                    group = []
        data += groups
    return pd.DataFrame(data, columns=["book_id", "author", "text"])


def visualize_groups(classes, groups):
    """
    Function that visualizes the partitions in the data made by the classes and existing groups.

    :param classes: Series corresponding to the target classes.
    :param groups: Series corresponding to the existing groups.
    """
    y, _ = classes.factorize()
    j = y[0]
    y_index = [0]
    for n, i in enumerate(y):
        if i != j:
            y_index.append(n)
        j = i
    y_index.append(len(y))

    g, _ = groups.factorize()
    j = g[0]
    g_index = [0]
    for n, i in enumerate(g):
        if i != j:
            g_index.append(n)
        j = i
    g_index.append(len(g))

    # Visualize dataset groups
    fig, ax = plt.subplots()
    ax.vlines(g_index, 0.5, 1, colors="darkorange", linestyles="solid", label="Groups", lw=1)
    ax.vlines(y_index, 0.5, 1, colors="darkblue", linestyles="dashed", label="Classes")
    ax.legend()
    ax.set_ylim(0, 1.5)
    ax.get_yaxis().set_visible(False)
    ax.set_xlabel("Sample index")
    return fig.show()


def plot_cv_indices(cv, X, y, group, n_splits, lw=10):
    """
    Create a sample plot for indices of a cross-validation object. Visualize how the cross-validation partitions the
    data.

    :param cv: cross_validation object.
    :param X: data to partition with cv.
    :param y: class labels to partition with.
    :param group: group labels to partition with.
    :param n_splits: number of folds.
    :param lw: line width.
    """
    cmap_cv = plt.cm.coolwarm
    cmap_data = plt.cm.Paired
    fig, ax = plt.subplots()
    # Generate the training/testing visualizations for each CV split
    for ii, (tr, tt) in enumerate(cv.split(X=X, y=y, groups=group)):
        # Fill in indices with the training/test groups
        indices = np.array([np.nan] * len(X))
        indices[tt] = 1
        indices[tr] = 0

        # Visualize the results
        ax.scatter(range(len(indices)), [ii + .5] * len(indices),
                   c=indices, marker='_', lw=lw, cmap=cmap_cv,
                   vmin=-.2, vmax=1.2)

    # Plot the data classes and groups at the end
    ax.scatter(range(len(X)), [ii + 1.5] * len(X),
               c=y, marker='_', lw=lw, cmap=cmap_data)

    ax.scatter(range(len(X)), [ii + 2.5] * len(X),
               c=group, marker='_', lw=lw, cmap=cmap_data)

    # Formatting
    yticklabels = list(range(n_splits)) + ['class', 'group']
    ax.set(yticks=np.arange(n_splits+2) + .5, yticklabels=yticklabels,
           xlabel='Sample index', ylabel="CV iteration",
           ylim=[n_splits+2.2, -.2])
    ax.set_title('{}'.format(type(cv).__name__), fontsize=15)
    return fig.show()


def plot_cm(confusion_matrix: np.array, class_names: list):
    """
    Function that creates a confusion matrix plot using the Wikipedia convention for the axis.

    :param confusion_matrix: confusion matrix that will be plotted
    :param class_names: labels of the classes

    Returns:
        - Plot of the Confusion Matrix
    """

    fig, ax = plt.subplots()
    plt.imshow(confusion_matrix, cmap=plt.get_cmap('cividis'))
    plt.colorbar()

    # We want to show all ticks...
    ax.set_xticks(np.arange(len(class_names)))
    ax.set_yticks(np.arange(len(class_names)))
    # ... and label them with the respective list entries
    ax.set_xticklabels(class_names)
    ax.set_yticklabels(class_names)

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")

    # Loop over data dimensions and create text annotations.
    for i in range(len(class_names)):
        for j in range(len(class_names)):
            ax.text(j, i, confusion_matrix[i, j], ha="center", va="center", color="w")

    ax.set_title("Confusion Matrix")
    plt.ylabel('Targets', fontweight='bold')
    plt.xlabel('Predictions', fontweight='bold')
    plt.ylim(top=len(class_names) - 0.5)  # adjust the top leaving bottom unchanged
    plt.ylim(bottom=-0.5)  # adjust the bottom leaving top unchanged
    plt.tight_layout()
    return plt.show()


def word_counter(text_list, top=None):
    """
    Function that creates the ranking of word frequencies in every string in a
    list of strings.

    :param text_list: list of strings.
    :param top: integer that indicates how many top ranks to display.

    Returns:
        - Series with the value counts of each top word.
    """
    split = [i.split() for i in text_list]
    words_in_df = list(itertools.chain.from_iterable(split))  # merges lists in list
    freq = pd.Series(words_in_df).value_counts()
    if top:
        return freq.head(top)
    else:
        return freq


def save_excel(dataframe, sheetname, filename="metrics"):
    """
    Function that saves the evaluation metrics into a excel file.

    :param dataframe: dataframe that contains the metrics.
    :param sheetname: name of the sheet containing the parameterization.
    :param filename: specifies the name of the xlsx file.

    """
    if not os.path.isfile("./outputs/{}.xlsx".format(filename)):
        mode = 'w'
    else:
        mode = 'a'
    writer = pd.ExcelWriter('./outputs/{}.xlsx'.format(filename), engine='openpyxl', mode=mode)
    dataframe.to_excel(writer, sheet_name=sheetname)
    writer.save()


def get_top_n_grams(corpus, top_k, n):
    """
    Function that receives a list of documents (corpus) extracts
        the top k most frequent n-grams for that corpus and plots the frequencies in a bar plot.

    :param corpus: list of texts
    :param top_k: int with the number of n-grams that we want to extract
    :param n: n gram type to be considered
             (if n=1 extracts unigrams, if n=2 extracts bigrams, ...)

    :return: Returns a sorted dataframe in which the first column
        contains the extracted ngrams and the second column contains
        the respective counts
    """
    vec = CountVectorizer(ngram_range=(n, n), max_features=2000).fit(corpus)

    bag_of_words = vec.transform(corpus)

    sum_words = bag_of_words.sum(axis=0)

    words_freq = []
    for word, idx in vec.vocabulary_.items():
        words_freq.append((word, sum_words[0, idx]))

    words_freq = sorted(words_freq, key=lambda x: x[1], reverse=True)
    top_df = pd.DataFrame(words_freq[:top_k])
    top_df.columns = ["Ngram", "Freq"]

    x_labels = top_df["Ngram"][:30]
    y_pos = np.arange(len(x_labels))
    values = top_df["Freq"][:30]
    plt.bar(y_pos, values, align='center', alpha=0.5)
    plt.xticks(y_pos, x_labels)
    plt.ylabel('Frequencies')
    plt.title('Words')
    plt.xticks(rotation=90)
    plt.show()
    return top_df


def model_selection(grids, X_train, y_train, X_test, y_test, grid_labels):
    """
        Function receives a list of GridSearch objects and fits them to the data. Returns the best parameters in each
        grid together with the respective macro recall score. Writes and excel file with information on the best
        parameters for each Pipeline and exports the Pipelines to pickle files.

        :param grids: list of GridSearch objects
        :param X_train: X train
        :param y_train: target train
        :param X_test: X test
        :param y_test: target test
        :param grid_labels: list with a label for each GridSearch object

        """
    print('Performing model optimizations...')
    for gs, label in zip(grids, grid_labels):
        print('\nEstimator: %s' % label)
        if os.path.isfile("./outputs/Pipelines.xlsx"):
            wb = load_workbook("./outputs/Pipelines.xlsx", read_only=True)
            if label in wb.sheetnames:
                print('Grid already fitted. Continue to next grid.')
                continue
        # Fit grid search
        gs.fit(X_train, y_train)
        # Best params
        print('Best params: %s' % gs.best_params_)
        # Best training data score
        print('Best training score: %.3f' % gs.best_score_)
        # Predict on test data with best params
        y_pred = gs.predict(X_test)
        # Test data score of model with best params
        print('Test set score for best params: %.3f ' % recall_score(y_test, y_pred, average="macro"))
        # Save Pipeline name and best parameters to excel
        grid_results = pd.DataFrame(gs.cv_results_)
        grid_results["best_index"] = pd.Series(np.zeros(grid_results.shape[0]))
        grid_results.loc[gs.best_index_, "best_index"] = 1
        save_excel(grid_results, label, "Pipelines")
        # Save fitted grid to pickle file
        dump(gs, "./outputs/Pipeline_{}.pkl".format(label), compress=1)
    print("Model Selection process finished")


def model_assessment_vis(path, labels):
    """
    Produces a bar plot for mean test score comparison together with error.
    :param path: path to excel file with each sheet representing a GridSearchCV results.
    :param labels: dictionary that contains grid labels (e.g. cv, knn) as keys and visualization labels as values
    (e.g. CountVectorizer, KNearestNeighbors).

    Returns: Plotly Visualization
    """
    xls = pd.ExcelFile(path, engine="openpyxl")
    dfs = [pd.read_excel(xls, i) for i in xls.sheet_names]

    update_dfs = []
    for i, df in enumerate(dfs):
        params = xls.sheet_names[i].split("_")
        df = df.loc[df.best_index == 1, ['mean_test_score', 'std_test_score']]
        df['model'] = labels[params[1]]
        df['encoding'] = labels[params[0]]
        update_dfs.append(df)

    xls.close()
    vis = pd.concat(update_dfs, ignore_index=True)
    vis_tfidf = vis.loc[vis.encoding == 'TF-IDF']
    vis_cv = vis.loc[vis.encoding == 'CountVectorizer']

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Bag of Words',
        x=vis_cv['model'], y=vis_cv['mean_test_score'],
        error_y=dict(type='data', array=vis_cv['std_test_score'])
    ))
    fig.add_trace(go.Bar(
        name='TF-IDF',
        x=vis_tfidf['model'], y=vis_tfidf['mean_test_score'],
        error_y=dict(type='data', array=vis_tfidf['std_test_score'])
    ))

    fig.update_layout(
        barmode='group',
        title="MODEL COMPARISON",
        xaxis_title="Model",
        yaxis_title="Recall average"
    )

    return pyo.plot(fig)


# class RepeatedStratifiedGroupKFold():
#     """
#     This ensures that excerpts belonging to the same book will always be in the same fold and that the proportion of
#     samples for each class is the same in all the folds.
#     https://github.com/scikit-learn/scikit-learn/issues/13621#issuecomment-557802602
#
#     :param n_splits: number of folds to produce (default = 5).
#     :param n_repeats: number of times each fold is used as validation (default = 1).
#     :param random_state: random state for reproducibility.
#     """
#
#     def __init__(self, n_splits=5, n_repeats=1, random_state=None):
#         self.n_splits = n_splits
#         self.n_repeats = n_repeats
#         self.random_state = random_state
#
#     # Implementation based on this kaggle kernel:
#     #    https://www.kaggle.com/jakubwasikowski/stratified-group-k-fold-cross-validation
#     def split(self, X, y=None, groups=None):
#         k = self.n_splits
#
#         def eval_y_counts_per_fold(y_counts, fold):
#             y_counts_per_fold[fold] += y_counts
#             std_per_label = []
#             for label in range(labels_num):
#                 label_std = np.std(
#                     [y_counts_per_fold[i][label] / y_distr[label] for i in range(k)]
#                 )
#                 std_per_label.append(label_std)
#             y_counts_per_fold[fold] -= y_counts
#             return np.mean(std_per_label)
#
#         rnd = check_random_state(self.random_state)
#         for repeat in range(self.n_repeats):
#             labels_num = np.max(y) + 1
#             y_counts_per_group = defaultdict(lambda: np.zeros(labels_num))
#             y_distr = Counter()
#             for label, g in zip(y, groups):
#                 y_counts_per_group[g][label] += 1
#                 y_distr[label] += 1
#
#             y_counts_per_fold = defaultdict(lambda: np.zeros(labels_num))
#             groups_per_fold = defaultdict(set)
#
#             groups_and_y_counts = list(y_counts_per_group.items())
#             rnd.shuffle(groups_and_y_counts)
#
#             for g, y_counts in sorted(groups_and_y_counts, key=lambda x: -np.std(x[1])):
#                 best_fold = None
#                 min_eval = None
#                 for i in range(k):
#                     fold_eval = eval_y_counts_per_fold(y_counts, i)
#                     if min_eval is None or fold_eval < min_eval:
#                         min_eval = fold_eval
#                         best_fold = i
#                 y_counts_per_fold[best_fold] += y_counts
#                 groups_per_fold[best_fold].add(g)
#
#             all_groups = set(groups)
#             for i in range(k):
#                 train_groups = all_groups - groups_per_fold[i]
#                 test_groups = groups_per_fold[i]
#
#                 train_indices = [i for i, g in enumerate(groups) if g in train_groups]
#                 test_indices = [i for i, g in enumerate(groups) if g in test_groups]
#
#                 yield train_indices, test_indices


class StratifiedGroupKFold(_BaseKFold):
    """Stratified K-Folds iterator variant with non-overlapping groups.

    This cross-validation object is a variation of StratifiedKFold that returns
    stratified folds with non-overlapping groups. The folds are made by
    preserving the percentage of samples for each class.

    The same group will not appear in two different folds (the number of
    distinct groups has to be at least equal to the number of folds).

    The difference between GroupKFold and StratifiedGroupKFold is that
    the former attempts to create balanced folds such that the number of
    distinct groups is approximately the same in each fold, whereas
    StratifiedGroupKFold attempts to create folds which preserve the
    percentage of samples for each class.

    Read more in the :ref:`User Guide <cross_validation>`.

    Parameters
    ----------
    n_splits : int, default=5
        Number of folds. Must be at least 2.

    shuffle : bool, default=False
        Whether to shuffle each class's samples before splitting into batches.
        Note that the samples within each split will not be shuffled.

    random_state : int or RandomState instance, default=None
        When `shuffle` is True, `random_state` affects the ordering of the
        indices, which controls the randomness of each fold for each class.
        Otherwise, leave `random_state` as `None`.
        Pass an int for reproducible output across multiple function calls.
        See :term:`Glossary <random_state>`.

    Examples
    --------
    >>> import numpy as np
    >>> from sklearn.model_selection import StratifiedGroupKFold
    >>> X = np.ones((17, 2))
    >>> y = np.array([0, 0, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    >>> groups = np.array([1, 1, 2, 2, 3, 3, 3, 4, 5, 5, 5, 5, 6, 6, 7, 8, 8])
    >>> cv = StratifiedGroupKFold(n_splits=3)
    >>> for train_idxs, test_idxs in cv.split(X, y, groups):
    ...     print("TRAIN:", groups[train_idxs])
    ...     print("      ", y[train_idxs])
    ...     print(" TEST:", groups[test_idxs])
    ...     print("      ", y[test_idxs])
    TRAIN: [2 2 4 5 5 5 5 6 6 7]
           [1 1 1 0 0 0 0 0 0 0]
     TEST: [1 1 3 3 3 8 8]
           [0 0 1 1 1 0 0]
    TRAIN: [1 1 3 3 3 4 5 5 5 5 8 8]
           [0 0 1 1 1 1 0 0 0 0 0 0]
     TEST: [2 2 6 6 7]
           [1 1 0 0 0]
    TRAIN: [1 1 2 2 3 3 3 6 6 7 8 8]
           [0 0 1 1 1 1 1 0 0 0 0 0]
     TEST: [4 5 5 5 5]
           [1 0 0 0 0]

    See also
    --------
    StratifiedKFold: Takes class information into account to build folds which
        retain class distributions (for binary or multiclass classification
        tasks).

    GroupKFold: K-fold iterator variant with non-overlapping groups.
    """

    def __init__(self, n_splits=5, shuffle=False, random_state=None):
        super().__init__(n_splits=n_splits, shuffle=shuffle,
                         random_state=random_state)

    # Implementation based on this kaggle kernel:
    # https://www.kaggle.com/jakubwasikowski/stratified-group-k-fold-cross-validation
    def _iter_test_indices(self, X, y, groups):
        labels_num = np.max(y) + 1
        y_counts_per_group = defaultdict(lambda: np.zeros(labels_num))
        y_distr = Counter()
        for label, group in zip(y, groups):
            y_counts_per_group[group][label] += 1
            y_distr[label] += 1

        y_counts_per_fold = defaultdict(lambda: np.zeros(labels_num))
        groups_per_fold = defaultdict(set)

        groups_and_y_counts = list(y_counts_per_group.items())
        rng = check_random_state(self.random_state)
        if self.shuffle:
            rng.shuffle(groups_and_y_counts)

        for group, y_counts in sorted(groups_and_y_counts,
                                      key=lambda x: -np.std(x[1])):
            best_fold = None
            min_eval = None
            for i in range(self.n_splits):
                y_counts_per_fold[i] += y_counts
                std_per_label = []
                for label in range(labels_num):
                    std_per_label.append(np.std(
                        [y_counts_per_fold[j][label] / y_distr[label]
                         for j in range(self.n_splits)]))
                y_counts_per_fold[i] -= y_counts
                fold_eval = np.mean(std_per_label)
                if min_eval is None or fold_eval < min_eval:
                    min_eval = fold_eval
                    best_fold = i
            y_counts_per_fold[best_fold] += y_counts
            groups_per_fold[best_fold].add(group)

        for i in range(self.n_splits):
            test_indices = [idx for idx, group in enumerate(groups)
                            if group in groups_per_fold[i]]
            yield test_indices


class RepeatedStratifiedGroupKFold(_RepeatedSplits):
    """Repeated Stratified K-Fold cross validator.

    Repeats Stratified K-Fold with non-overlapping groups n times with
    different randomization in each repetition.

    Read more in the :ref:`User Guide <cross_validation>`.

    Parameters
    ----------
    n_splits : int, default=5
        Number of folds. Must be at least 2.

    n_repeats : int, default=10
        Number of times cross-validator needs to be repeated.

    random_state : int or RandomState instance, default=None
        Controls the generation of the random states for each repetition.
        Pass an int for reproducible output across multiple function calls.
        See :term:`Glossary <random_state>`.

    Examples
    --------
    >>> import numpy as np
    >>> from sklearn.model_selection import RepeatedStratifiedGroupKFold
    >>> X = np.ones((17, 2))
    >>> y = np.array([0, 0, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    >>> groups = np.array([1, 1, 2, 2, 3, 3, 3, 4, 5, 5, 5, 5, 6, 6, 7, 8, 8])
    >>> cv = RepeatedStratifiedGroupKFold(n_splits=2, n_repeats=2,
    ...                                   random_state=36851234)
    >>> for train_index, test_index in cv.split(X, y, groups):
    ...     print("TRAIN:", groups[train_idxs])
    ...     print("      ", y[train_idxs])
    ...     print(" TEST:", groups[test_idxs])
    ...     print("      ", y[test_idxs])
    TRAIN: [2 2 4 5 5 5 5 8 8]
           [1 1 1 0 0 0 0 0 0]
     TEST: [1 1 3 3 3 6 6 7]
           [0 0 1 1 1 0 0 0]
    TRAIN: [1 1 3 3 3 6 6 7]
           [0 0 1 1 1 0 0 0]
     TEST: [2 2 4 5 5 5 5 8 8]
           [1 1 1 0 0 0 0 0 0]
    TRAIN: [3 3 3 4 7 8 8]
           [1 1 1 1 0 0 0]
     TEST: [1 1 2 2 5 5 5 5 6 6]
           [0 0 1 1 0 0 0 0 0 0]
    TRAIN: [1 1 2 2 5 5 5 5 6 6]
           [0 0 1 1 0 0 0 0 0 0]
     TEST: [3 3 3 4 7 8 8]
           [1 1 1 1 0 0 0]

    Notes
    -----
    Randomized CV splitters may return different results for each call of
    split. You can make the results identical by setting `random_state`
    to an integer.

    See also
    --------
    RepeatedStratifiedKFold: Repeats Stratified K-Fold n times.
    """

    def __init__(self, n_splits=5, n_repeats=10, random_state=None):
        super().__init__(StratifiedGroupKFold, n_splits=n_splits,
                         n_repeats=n_repeats, random_state=random_state)